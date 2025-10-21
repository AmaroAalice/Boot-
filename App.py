

import os
import re
import openpyxl
import pyautogui
from urllib.parse import quote
import webbrowser
from time import sleep
import datetime



pasta = r'Y:\4. Pilar Entrega\4.0 Satisfação do Cliente\4.1 Nível de Serviço ao Cliente\Rating'
arquivos =  [os.path.join(pasta, f) for f in os.listdir(pasta) if f.endswith('.xlsx')]

# Verifica se encontrou algum arquivo
if not arquivos:
    raise FileNotFoundError(f"Nenhum arquivo .xlsx encontrado em {pasta}")

arquivo_mais_recente = max(arquivos, key=os.path.getctime)


print(f"📂 Arquivo mais recente encontrado: {arquivo_mais_recente}")

woorkbook = openpyxl.load_workbook(arquivo_mais_recente)
pagina_clientes = woorkbook['RANTING'] if 'RANTING' in woorkbook.sheetnames else woorkbook.active


# Abre o WhatsApp Web
webbrowser.open('https://web.whatsapp.com/')    
input("Escaneie o QR Code do WhatsApp e pressione Enter para continuar...")


for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    # Pdv,Nome Pdv ,contato,data chamada,data atendimento,motivo
    Pdv = linha[0]
    Nome_Pdv = linha[1]
    contato = linha[2]
    data_chamada = linha[5] if len(linha) > 5 else None
    data_atendimento = linha[6] if len(linha) > 6 else None
    motivo = linha[8] if len(linha) > 8 else None


    if not contato:
        continue

    # sanitize contato: keep digits and optional leading +
    contato_str = str(contato or "").strip()
    contato_digits = re.sub(r'[^\d+]', '', contato_str)

    # remove leading zeros (user should provide international code)
    if contato_digits.startswith('0'):
        contato_digits = contato_digits.lstrip('0')

    # format date safely
    if isinstance(data_chamada, (datetime.date, datetime.datetime)):
        data_str = data_chamada.strftime('%d/%m/%y')
    else:
        data_str = str(data_chamada or "")

         # Mensagem personalizada
    mensagem = ( f"Olá {Nome_Pdv}, tudo bem? 😊\n"
        f"Gostaria de entender o motivo da sua Avaliação do dia {data_str} ser '{motivo}'.\n"
        "Podemos agendar uma conversa para melhorarmos sua experiência?"
    )
    link_mensagem_whatsapp = f"https://api.whatsapp.com/send?phone={contato_digits}&text={quote(mensagem)}"

    print(link_mensagem_whatsapp)

    try:
        webbrowser.open(link_mensagem_whatsapp)
        sleep(10)
        pyautogui.press('enter')
        sleep(3)
        pyautogui.hotkey('ctrl', 'w')
        sleep(3)
        print(f"✅ Mensagem enviada para {Nome_Pdv}")
    except Exception as e:
        print(f"⚠️ Erro ao enviar mensagem para {Nome_Pdv}: {e}")
        with open('erros.csv', 'a', encoding='utf-8') as arquivo_erro:
            arquivo_erro.write(f"{Pdv},{Nome_Pdv},{contato}\n")

print("🏁 Processo finalizado com sucesso!")
