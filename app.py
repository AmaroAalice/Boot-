import os
import re
import sys
import openpyxl
from urllib.parse import quote
from time import sleep
import datetime
import platform

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from dotenv import load_dotenv
load_dotenv()

PLANILHAS_DIR = os.getenv("PLANILHAS_DIR")
os.makedirs(PLANILHAS_DIR, exist_ok=True)

print(f"O programa, por padrão, utiliza o diretório '{PLANILHAS_DIR}' para ler as planilhas.\n")
pasta_input = input("Pressione Enter para manter ou digite outro diretório: ").strip()
if pasta_input:
    PLANILHAS_DIR = pasta_input

arquivos = [os.path.join(PLANILHAS_DIR, f) for f in os.listdir(PLANILHAS_DIR) if f.endswith('.xlsx')]
if not arquivos:
    raise FileNotFoundError(f"Nenhum arquivo .xlsx encontrado em {PLANILHAS_DIR}")

arquivo_mais_recente = max(arquivos, key=os.path.getctime)
print(f"📂 Arquivo mais recente encontrado: {arquivo_mais_recente}")

workbook = openpyxl.load_workbook(arquivo_mais_recente)
pagina_clientes = workbook['RANTING'] if 'RANTING' in workbook.sheetnames else workbook.active

chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--disable-software-rasterizer")
chrome_options.add_argument("--remote-debugging-port=9222")

SO = platform.system().lower()

if getattr(sys, 'frozen', False):
    if SO.startswith("win"):
        chromedriver_path = os.path.join(sys._MEIPASS, 'chrome_drivers', 'chromedriver.exe')
    else:
        chromedriver_path = os.path.join(sys._MEIPASS, 'chrome_drivers', 'chromedriver-linux')
else:
    if SO.startswith("win"):
        chromedriver_path = os.path.join('chrome_drivers', 'chromedriver.exe')
    else:
        chromedriver_path = os.path.join('chrome_drivers', 'chromedriver-linux')

service = Service(chromedriver_path)
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.get("https://web.whatsapp.com/")
input("📱 Após escanear o QR Code, pressione Enter para continuar...")

for idx, linha in enumerate(pagina_clientes.iter_rows(min_row=2, values_only=True), start=2):
    row_num = idx
    Pdv = linha[0]
    Nome_Pdv = linha[1]
    contato = linha[2]
    data_chamada = linha[5]
    data_atendimento = linha[6]
    motivo = linha[8]

    if not contato or data_atendimento:
        continue

    contato_str = str(contato).strip()
    contato_digits = re.sub(r'[^\d+]', '', contato_str)
    if contato_digits.startswith('0'):
        contato_digits = contato_digits.lstrip('0')

    if isinstance(data_chamada, (datetime.date, datetime.datetime)):
        data_str = data_chamada.strftime('%d/%m/%y')
    else:
        data_str = str(data_chamada or "")

    if not motivo:
        motivo = "Sem motivo especificado"

    mensagem = (
        f"Olá {Nome_Pdv}, tudo bem?\n"
        f"Sou A Alice da Ambev,Gostaria de entender o motivo da sua avaliação do dia {data_str} ser '{motivo}'.\n"
        "Podemos agendar uma conversa para melhorarmos sua experiência?"
    )

    link_whatsapp = f"https://web.whatsapp.com/send?phone={contato_digits}"
    driver.get(link_whatsapp)

    try:
        # *** CORREÇÃO DA CAIXA DE MENSAGEM – SEM MEXER NO RESTO ***
        caixa_msg = WebDriverWait(driver, 25).until(
            EC.presence_of_element_located((
                By.XPATH,
                "//footer//p[contains(@class,'selectable-text')]"
            ))
        )

        sleep(8)
        caixa_msg.click()
        sleep(8)

        for linha_msg in mensagem.split("\n"):
            caixa_msg.send_keys(linha_msg)
            caixa_msg.send_keys(Keys.SHIFT + Keys.ENTER)

        caixa_msg.send_keys(Keys.ENTER)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.message-out"))
        )

        sleep(1)
        data_atendimento = datetime.datetime.now().date().strftime('%d/%m/%Y')
        pagina_clientes.cell(row=row_num, column=6, value=data_atendimento)
        workbook.save(arquivo_mais_recente)

        print(f"✅ Mensagem enviada para {Nome_Pdv}")

    except Exception as e:
        print(f"⚠️ Erro ao enviar para {Nome_Pdv}: {e}")
        with open('erros.csv', 'a', encoding='utf-8') as arquivo_erro:
            arquivo_erro.write(f"@{e} - {Pdv}, {Nome_Pdv}, {contato}\n")

driver.quit()
print("🏁 Processo finalizado com sucesso!")
